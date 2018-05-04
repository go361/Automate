#! python3
# -*- coding: utf-8 -*-
# @date: 2018/5/4 16:14
# @name: ch12multiplicationTable
# @author：Go361
"""
Input a Number, The program will give you a
Table of multiplication:
"""
import openpyxl
from openpyxl.cell.cell import (get_column_letter, column_index_from_string
								)
N = int(input("Please Input a Number:"))
fileName = "Table of multiplication"
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
sheet.title = fileName

for n in range(1, N + 1):
	sheet.row_dimensions[n].height = 20
	sheet.column_dimensions[get_column_letter(n)].width = 10
	sheet.cell(row = 1, column = n + 1).value = n
	sheet.cell(row = n + 1, column = 1).value = n
for row in range(1, N + 1):
	for col in range(1, N + 1):
		sheet.cell(row = row + 1, column = col + 1).value = row * col

wb.save(fileName + ".xlsx")